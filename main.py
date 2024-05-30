#FILENAME:main.py
#AUTHOR:Jonathan Shambaugh
#PURPOSE: A CLI tool that consolidates useful python tools for EDC.
#NOTES: See the github repository for more info.
#VERSION: v0.0.0
#START DATE: 17 Oct 22


from sys import executable, version
from platform import system, platform
from subprocess import check_call, check_output
from os import system, access, getcwd, listdir, environ, path, R_OK, W_OK, X_OK
from csv import reader
from time import perf_counter
from concurrent.futures import ThreadPoolExecutor
from typing import Optional
from shutil import copyfile
import json
import socket
import atexit


#defining program constants
V = "v0.0.0"


T1 = perf_counter()


OP_SYS = platform().split("-")[0]


#installs libraries using the command line
def install_lib(lib):
    print(f"\nInstalling {lib}...")
    try:
        # implement pip as a subprocess:
        check_call([executable, '-m', 'pip', 'install', lib])
    except:
        check_call([executable, '-m', 'pip3', 'install', lib])

    # process output with an API in the subprocess module:
    requests = check_output([executable, '-m', 'pip','freeze'])
    installed_packages = [r.decode().split('==')[0] for r in requests.split()]

    logger.info('Installed: %s',installed_packages)


#check if picologging library is installed, if not, install it
if OP_SYS == "Mac":
    import logging as logging
else:
    try:
        import picologging as logging
    except ModuleNotFoundError:
        print(f"Picologging library is not installed.")
        install_lib("picologging")
        import picologging as logging

#configure logger
logger = logging.getLogger(__name__)
logging.basicConfig(filename="data.log", level=logging.INFO, format='%(asctime)s | %(name)s - %(levelname)s - %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')


#check if openpyxl library is installed, if not, install it
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print(f"Openpyxl library is not installed.")
    install_lib("Openpyxl")
    from openpyxl import load_workbook

#check if requests library is installed, if not, install it
try:
    from requests import get
except ModuleNotFoundError:
    print(f"Requests library is not installed.")
    install_lib("requests")
    from requests import get

#check if alive-progress library is installed, if not, install it
try:
    from tqdm import tqdm
except ModuleNotFoundError:
    print(f"tqdm library is not installed.")
    install_lib("tqdm")
    from tqdm import tqdm

#check if colorama library is installed, if not, install it
try:
    from colorama import *
except ModuleNotFoundError:
    print(f"Colorama library is not installed.")
    install_lib("colorama")
    from colorama import *


#define config class for easier attribute access and control
class Config:
    
    def __init__(self):
        config_exists = path.isfile("config.json")
        if not config_exists:
            k = open("config.json", "x")
            k.close()
            with open("config.json", "w") as h:
                default_config = {'User': environ.get('USERNAME'), 'Log Level': 'INFO', 'Reset': True,'Input Directory': './', 'Template Directory': './', 'Output Directory': './'}
                write_config = json.dumps(default_config)
                h.write(write_config)
                h.close()
        with open("config.json", "r") as file:
            config_data = json.load(file)
            
        print(config_data)            
        file.close()
        self._user = config_data['User']
        self._log_level = config_data['Log Level']
        self._reset = config_data['Reset']
        self._input_dir = config_data['Input Directory']
        self._output_dir = config_data['Output Directory']
        self._template_dir = config_data['Template Directory']
        match self._log_level:
            case 'INFO':
                logger.setLevel(logging.INFO)
            case 'DEBUG':
                logger.setLevel(logging.DEBUG)
            case 'WARN':
                logger.setLevel(logging.WARN)
            case 'ERROR':
                logger.setLevel(logging.ERROR)
            case 'CRITICAL':
                logger.setLevel(logging.CRITICAL)

    
    def get_log_level(self) -> str:
        return self._log_level
    
    
    def get_input_dir(self) -> str:
        return self._input_dir
    
    
    def get_output_dir(self) -> str:
        return self._output_dir
    
    
    def get_temp_dir(self) -> str:
        return self._template_dir
    

    def change_log_level(level) -> None:
        match level:
            case 'INFO':
                logger.setLevel(logging.INFO)
            case 'DEBUG':
                logger.setLevel(logging.DEBUG)
            case 'WARN':
                logger.setLevel(logging.WARN)
            case 'ERROR':
                logger.setLevel(logging.ERROR)
            case 'CRITICAL':
                logger.setLevel(logging.CRITICAL)
        return None



#function for required tasks before program exit
@atexit.register
def exit_handler() -> None:
    time_elapsed = round((perf_counter() - T1), 3)
    logger.debug('Execution time: %s sec(s)', time_elapsed)


#function for tasks that need to be completed during startup
def preamble() -> None:
    system('color')
    print(f"Events Layout Import Tool")
    owner = "jkernal"
    repo = "EDCTools"
    print("Checking for updates...", end="",flush=True)
    try:
        response = get(f"https://api.github.com/repos/{owner}/{repo}/releases/latest")
        print(response.json())
        if V != response.json()["tag_name"]:
            print("[DONE]")
            print(f"***Warning: There is a new release of this tool.***")
    except:
        print("[FAILED]")
        print("***Warning: Could not connect to repository. Version check failed.***")
    print("""                                                 
   ↖↗→→→→→→→→→→→→→→↗→→→→→→→→→→↙          ↖↓→↓←←←←↙↘↘←       
   ↖→↖↖↖↖↖↖↖↖↖↖↖↖↖↖↑↖↖↖↖↖↖↖↖↖↖↙↙  ←    →↙←↖↖↖←↓↓↙↖↖↖←↙↘←    
   ↖→↖↖↖↖↖↖↖↖↖↖↖↖↖↖↑↖↖↖↖↖↖↖↖↖↖←↙↙↓←↓↓↘↓↖↖↙→→→→→→→→↖↖↖↙↓↖    
   ↖→↖↖↖↖↖↙↓↓↓↓↓↓↓↓↑↖↖↖↖↖↙↙←↖↖↖↖↖↖↖↖↓↙↖←↑↑↑→→→→→→↓↙←↓↙      
   ↖→↖↖↖↖↖↓←       ↑↖↖↖↖←↑ ↖↓↓←↖↖↖↖↙↙↖↙↑↑↑↑↘      ↖↙←       
   ↖→↖↖↖↖↖↓↓←←←←←←↙↑↖↖↖↖←↑    ↓↙↖↖↖↖↓↙↑↑↑↑   ↖↓             
   ↖→↖↖↖↖↖↖↖↖↖↖↖↖↖↖↑↖↖↖↖←↑     ↘↓↖↖↖↖↖↖↑↑↖   ↓↗             
   ↖→↖↖↖↖↖↖↖↖↖↖↖↖↖↖↑↖↖↖↖←↑     →↓↖↖↖↖↖↖↑→↖  ↓↑↑↘            
   ↖→↖↖↖↖↖↖↖↖↖↖↖↖↖↖↑↖↖↖↖←↑     ↓↙↖↖↖↑↓←←←↗  ↖↑↗             
   ↖→↖↖↖↖↖↙↙      ↖↑↖↖↖↖←↑    ↓↓↖↖↖↖↑←←←←←↑                 
   ↖→↖↖↖↖↖↙←       ↑↖↖↖↖←↑  ↘↘↖↖↖↖↖↖↖↙↙↙←←←←←↑↘↓↘↗↘↙↓←      
   ↖→↖↖↖↖↖←↙↙↙↙↙↙↙↓↑↖↖↖↖↖↙↙←↖↖↖↖↖↓↓↖↓↑←↖↖↙←←←←←↙↖↖↖↖↖↙↙     
   ↖→↖↖↖↖↖↖↖↖↖↖↖↖↖↖↑↖↖↖↖↖↖↖↖↖↖↖←←  ←← ↓↘↖↖↖↖↙↙↙←↖↖↖↖↖↖↙→↖   
   ↖→↖↖↖↖↘↓↖↖↖↖↓←↖↖↑↖↖→↖↖↖↙↗↖↖↖↙↙↖      ↓↙↓↙↖↖↖↖↖↖←↙↙↓↙     
    ↓↓↓↘↗↘↙↓↙↙↓→↗↓↘→→↓↗↘→↘↗→↙→←             ←↙↓↘↓↙←         
        ← ↓↗↙ ↙↖↘↗   ↓ ↑→ ←↙→                               
                       ←   ↖←   """)
    print("Type 'help' for a list of commands.")
    return None
#end of preamble


#check permissions on files needed
def perm_check(locs):
    file_names = ["template", "output", "input"]
    access_type = ["read", "write", "execute"]
    for i in range(len(locs)):
        permissions = [access(locs[i], R_OK), access(locs[i], W_OK), access(locs[i], X_OK)]
        for j in range(len(permissions)):
            if not permissions[j]:
                print(f"The script does not have {access_type[j]} access to the {file_names[i]} file. Make sure the file is closed and permissions are set.")
            else:
                #print(f"\u001b[1m\u001b[31;1mThe script does have {access_type[j]} access to the {file_names[i]} file. Make sure the file is closed and permissions are set.")
                continue
        continue
    return None


#Tool for extracting address comments dynamically 
def EventsTool(Config_obj: Config, acronym: str) -> None:
    
    
    #gets address that need comments
    def get_address_array_from_temp(sheet) -> list:
        array = []
        for i in range(sheet.max_row):
            array.append([sheet.cell(i+3,2).value, i + 3])
        return array
    
    
    #gets all address that have comments
    def get_address_comment_array_from_input(input_file_path) -> list:
        try:
            array = list(reader(open(input_file_path, encoding= "ISO8859")))
        except PermissionError:
            print(f"Error: Could not access input file.")
        except:
            print(f"""An error occurred while reading the Toyopuc Comment file.\n\n
                Possible Causes:\n
                -Too many fields in the file (max 131072)
                -The data is not supported under 'ISO8859' encoding
                -The file is in use""")
        return array


    #find file locations
    cwrd = getcwd()
    file_locs = dict()
    if Config_obj.get_input_dir() == "./":
        file_locs.update({"input": cwrd + "\\input"})
    else:
        file_locs.update({"input": Config_obj.get_input_dir()})
    if Config_obj.get_temp_dir() == "./":
        file_locs.update({"template": cwrd + "\\template"})
    else:
        file_locs.update({"template": Config_obj.get_temp_dir()})
    if Config_obj.get_output_dir() == "./":
        file_locs.update({"output": cwrd + "\\output"})
    else:
        file_locs.update({"output": Config_obj.get_output_dir()})
    
    #define path variables for files
    in_file_path = file_locs["input"] + "\\" + listdir(file_locs["input"])[0]
    temp_file_path = file_locs["template"] + "\\" + listdir(file_locs["template"])[0]
    out_file_path = file_locs["output"] + "\\" + acronym + ".xlsx"
    copyfile(temp_file_path, out_file_path)

    #check permissions on files
    perm_check([temp_file_path, out_file_path, in_file_path])

    #open output workbook and worksheet
    wb = load_workbook(filename=out_file_path)
    ws = wb["Import Cheat Sheet"]

    #get address that need comments from template and get addresses with comments from input in separate threads
    with ThreadPoolExecutor() as executor:
        f1 = executor.submit(get_address_array_from_temp, ws)
        f2 = executor.submit(get_address_comment_array_from_input, in_file_path)
    #wait for results from both threads
    address_array = f1.result()
    address_comment_array = f2.result()
    #close executor
    executor.shutdown()

    match_count, address_array_len = 0, len(address_array)
    
    print(f"\nWorking on it...")

    #loop through the addresses and compare to the array with comments
    for i in tqdm(range(address_array_len)):
        for address in address_comment_array:
            if address_array[i][0] == address[0]:
                ws.cell(row=address_array[i][1], column=6).value = address[0]
                ws.cell(row=address_array[i][1], column=7).value = address[1]
                match_count+=1
            elif address[0][:4] == "P1-X":
                if address_array[i][0] == address[0][3:]:
                    ws.cell(row=address_array[i][1], column=6).value = address[0][3:]
                    ws.cell(row=address_array[i][1], column=7).value = address[1]
                    match_count+=1
                else:
                    continue
            elif address[0][:4] == "P2-D":
                if address_array[i][0] == address[0][3:]:
                    ws.cell(row=address_array[i][1], column=6).value = address[0][3:]
                    ws.cell(row=address_array[i][1], column=7).value = address[1]
                    match_count+=1
                else:
                    continue
    
    #save changes to the output file
    wb.save(out_file_path)
    wb.close()
    
    #display stats and warning if needed
    print("\nDone.", flush=True)
    print(f"\nNumber of comments found:" + str(match_count))
    if match_count == 0:
        print(f"***No matches were found. Make sure your input and template files are correct***")

    return None
    #end of EventsTool


def validate_ipv4_address(ip_address: str, private_only: Optional[bool] = True) -> bool:
    #always accepts loopback address
    if ip_address == "127.0.0.1":
        return True
    octets = ip_address.split(".")
    if len(octets) != 4:
        return False
    for octet in octets:
        if not octet.isdigit():
            return False
        n = int(octet)
        if n < 0 or n > 255:
            return False
        else:
            continue
    if private_only:
        first_octet = int(octets[0])
        second_octet = int(octets[1])
        if first_octet == 10:
            pass
        elif first_octet == 172 and second_octet > 15 and second_octet < 32:
            pass
        elif first_octet == 192 and second_octet == 168:
            pass
        else:
            return False
    return True


def check_ports(ip_address: str, timeout: Optional[int] = 1) -> list:
    is_ipv4 = validate_ipv4_address(ip_address)
    if not is_ipv4:
        raise ValueError("Needs to be a valid IP address.")
    open_ports = []
    print(f"Attempting to connect to {ip_address}")
    for port in tqdm(range(7000, 7008)):
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)  # Set timeout to 1 second
        result = sock.connect_ex((ip_address, port))
        if result == 0:
            open_ports.append(port)
        sock.close()
    return open_ports


def main():
    config = Config()
    logger.debug('Python %s', version)
    logger.debug('%s', platform())
    preamble()
    user_input = ''
    print('Enter command:')
    while user_input != 'q':
        user_input = input('./>').lower()
        logger.info('Input entered: %s', user_input)
        command = user_input.split(" ")
        match command[0]:
            case 'help':
                print(" Command   |    Argument1   |    Argument2      ")
                print("eventstool       acronym           None         ")
                print("checkports     ip address      timeout(optional)")
                print("     q to quit")
            case 'eventstool':
                EventsTool(config, "FSMB")
            case 'checkports':
                try:
                    open_ports = check_ports(command[1])
                    if len(open_ports) > 0:
                        print("Ports open:")
                        for port in open_ports:
                            print(port)
                    else:
                        print("Unable to connect to any ports.")
                except ValueError:
                    print(f"{command[1]} is not a valid IP address.")


def print_argument(func):
    def wrapper(the_number):
        print("Argument for", func.__name__, "is", the_number)
        return func(the_number)
    return wrapper


if __name__ == '__main__':
    main()
